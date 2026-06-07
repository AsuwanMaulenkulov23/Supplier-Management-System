"""
Слой работы с БД (SQLite, стандартная библиотека — без внешних ORM).

Содержит:
  * описание схемы;
  * инициализацию БД;
  * засев (seed) данных из приложенных Excel-файлов;
  * создание тестовых пользователей для трёх ролей.

БД хранится в одном файле app.db рядом с проектом. Чтобы пересоздать
данные с нуля — просто удалите app.db и перезапустите приложение
(или запустите `python seed.py --reset`).
"""

import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "app.db")
SUPPLIERS_XLSX = os.path.join(BASE_DIR, "data", "suppliers.xlsx")
ORDERS_XLSX = os.path.join(BASE_DIR, "data", "orders.xlsx")

# Тестовые аккаунты (роль -> логин/пароль). Дублируются в README.
TEST_USERS = [
    ("buyer", "BUYER", "buyer", "Закупщик"),
    ("manager", "MANAGER", "manager", "Менеджер продаж"),
    ("director", "DIRECTOR", "director", "Руководитель"),
]

SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    username      TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role          TEXT NOT NULL,          -- buyer | manager | director
    full_name     TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS suppliers (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    fruit          TEXT NOT NULL,
    purchase_price REAL NOT NULL,         -- цена закупки, ₸/кг
    stock          REAL NOT NULL,         -- остаток на складе, кг
    name           TEXT NOT NULL,         -- наименование поставщика
    delivery_days  INTEGER NOT NULL       -- срок поставки, дней
);

CREATE TABLE IF NOT EXISTS orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    restaurant      TEXT NOT NULL,
    fruit           TEXT NOT NULL,
    quantity        REAL NOT NULL,        -- кг
    sale_price      REAL,                 -- цена продажи, ₸/кг (считает алгоритм)
    order_sum       REAL,                 -- сумма заказа, ₸ (считает алгоритм)
    order_date      TEXT NOT NULL,        -- YYYY-MM-DD
    client_due_date TEXT NOT NULL,        -- YYYY-MM-DD (срок поставки клиенту)
    status          TEXT NOT NULL,        -- Подобран поставщик | Требует ручного разбора
    supplier_id     INTEGER,              -- выбранный поставщик
    purchase_price  REAL,                 -- цена закупки выбранного поставщика (для прибыли)
    reason          TEXT,                 -- почему ушёл в ручной разбор
    FOREIGN KEY (supplier_id) REFERENCES suppliers (id)
);

-- История изменений (бонус из ТЗ): кто и когда менял запись.
CREATE TABLE IF NOT EXISTS audit_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    ts          TEXT NOT NULL,
    username    TEXT NOT NULL,
    entity      TEXT NOT NULL,            -- supplier | order
    entity_id   INTEGER,
    action      TEXT NOT NULL,            -- create | update | delete
    details     TEXT
);
"""


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _parse_date(value) -> str:
    """Excel может отдать дату как datetime или как строку '01.05.2026'. Нормализуем в ISO."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):  # datetime.date
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Не удалось разобрать дату: {value!r}")


def already_seeded(conn) -> bool:
    row = conn.execute("SELECT COUNT(*) AS c FROM suppliers").fetchone()
    return row["c"] > 0


def seed_users(conn):
    for username, password, role, full_name in TEST_USERS:
        conn.execute(
            "INSERT OR IGNORE INTO users (username, password_hash, role, full_name) "
            "VALUES (?, ?, ?, ?)",
            (username, generate_password_hash(password), role, full_name),
        )


def seed_suppliers(conn):
    wb = load_workbook(SUPPLIERS_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for r in rows:
        # № | Наименование(фрукт) | Цена закупки | Кол-во на складе | Поставщик | Срок поставки
        if r[0] is None:
            continue
        conn.execute(
            "INSERT INTO suppliers (fruit, purchase_price, stock, name, delivery_days) "
            "VALUES (?, ?, ?, ?, ?)",
            (str(r[1]).strip(), float(r[2]), float(r[3]), str(r[4]).strip(), int(r[5])),
        )
    wb.close()


def seed_orders(conn):
    """Заказы засеваются «как есть» и сразу прогоняются через алгоритм подбора.

    Импорт алгоритма делаем внутри функции, чтобы избежать кольцевых импортов.
    """
    from datetime import date as _date
    from algorithm import select_supplier, STATUS_MATCHED

    wb = load_workbook(ORDERS_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    for r in rows:
        # № | Ресторан | Фрукт | Кол-во | Цена продажи | Сумма | Дата заказа | Срок поставки
        if r[0] is None:
            continue
        restaurant = str(r[1]).strip()
        fruit = str(r[2]).strip()
        quantity = float(r[3])
        order_date = _parse_date(r[6])
        client_due = _parse_date(r[7])

        # Текущие (живые) остатки поставщиков, чтобы заказы списывали склад
        # последовательно, как в реальном времени.
        suppliers = [
            dict(s) for s in conn.execute("SELECT * FROM suppliers").fetchall()
        ]

        result = select_supplier(
            {
                "fruit": fruit,
                "quantity": quantity,
                "order_date": _date.fromisoformat(order_date),
                "client_due_date": _date.fromisoformat(client_due),
            },
            suppliers,
        )

        conn.execute(
            "INSERT INTO orders (restaurant, fruit, quantity, sale_price, order_sum, "
            "order_date, client_due_date, status, supplier_id, purchase_price, reason) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                restaurant,
                fruit,
                quantity,
                result["sale_price"],
                result["order_sum"],
                order_date,
                client_due,
                result["status"],
                result["supplier_id"],
                result["purchase_price"],
                result["reason"],
            ),
        )

        # Правило 4: списываем склад выбранного поставщика.
        if result["status"] == STATUS_MATCHED:
            conn.execute(
                "UPDATE suppliers SET stock = stock - ? WHERE id = ?",
                (quantity, result["supplier_id"]),
            )


def init_db(reset: bool = False):
    if reset and os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = get_conn()
    conn.executescript(SCHEMA)
    seed_users(conn)
    if not already_seeded(conn):
        seed_suppliers(conn)
        seed_orders(conn)
    conn.commit()
    conn.close()


if __name__ == "__main__":
    import sys

    reset = "--reset" in sys.argv
    init_db(reset=reset)
    print(
        "База инициализирована" + (" (пересоздана)" if reset else "") + f": {DB_PATH}"
    )
